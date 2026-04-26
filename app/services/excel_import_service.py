"""
Excel import service for TAT inventory exports.

Column mapping strategy
-----------------------
Columns are resolved via a priority-ordered list of case-insensitive aliases.
The first alias found in the sheet header wins.

Logical fields and their known aliases:
  item_name     → "Item Name", "ItemName", "Description", ...
  serial_number → "Serial Number", "SerialNumber", "SN", ...
  part_number   → "Part Number", "PartNumber", "Part#", "PN", ...
  part_desc     → "PartDescription", "Part Description", ...  [TAT-specific]
  status        → "System Status", "Status", "Estado", ...
  location      → "Rig Location", "Location", "Ubicación", ...
  size          → "Size", "Tamaño"  (appended to item_name unless "0")
  internal_code → "Internal Code", "Code", "Código Interno", ...
  series        → "Series", "Serie"  [TAT-specific, stored in internal_code
                   only when serial_number is absent]

Real TAT status values → AssetStatus mapping
--------------------------------------------
  "In"         → available   (in warehouse)
  "Dirty"      → maintenance (needs cleaning)
  "Field"      → in_use      (deployed)
  "In Service" → in_use      (actively in service)

Transaction strategy
--------------------
Each row uses an individual SAVEPOINT so a bad row never rolls back the
rest of the import. A single db.commit() persists all successful rows.
"""

import re
from io import BytesIO
from typing import Any

import openpyxl
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models.asset import Asset, AssetStatus
from app.models.part import Part
from app.schemas.excel_import import (
    ColumnMapping, ImportResult, ImportRowError,
)


# ── Constants ──────────────────────────────────────────────────────────────────

MAX_ERRORS_PER_IMPORT = 500
MAX_ROWS_PER_IMPORT   = 10_000
MAX_FIELD_LEN         = 255

# Size values that are TAT placeholders and should be ignored
SIZE_IGNORE_VALUES = {"0", "0.0", "none", "n/a", "na", "-"}

# ── Column alias registry ──────────────────────────────────────────────────────

COLUMN_ALIASES: dict[str, list[str]] = {
    "item_name": [
        "item name", "itemname", "item_name",
        "description", "descripcion", "descripción",
        "equipment name", "nombre", "nombre equipo",
    ],
    "serial_number": [
        "serial number", "serialnumber", "serial_number",
        "serial no", "serial no.", "serial#", "sn",
        "numero de serie", "nro serie", "nro. serie",
    ],
    "part_number": [
        "part number", "partnumber", "part_number",
        "part#", "part no", "part no.",
        "numero de parte", "nro parte", "nro. parte", "pn",
    ],
    "part_desc": [
        # TAT exports PartDescription as a separate column
        "partdescription", "part description", "part_description",
        "partdesc", "part desc",
    ],
    "status": [
        "system status", "systemstatus", "system_status",
        "status", "asset status", "estado", "estado equipo",
    ],
    "location": [
        "rig location", "riglocation", "rig_location",
        "location", "ubicacion", "ubicación", "rig loc",
    ],
    "size": ["size", "tamaño", "tamano"],
    "internal_code": [
        "internal code", "internalcode", "internal_code",
        "code", "asset code", "codigo interno", "código interno",
        "int code", "int. code",
    ],
    "series": [
        # TAT-specific: equipment series/family identifier
        "series", "serie",
    ],
}

# ── Status normalisation ───────────────────────────────────────────────────────
# Maps lowercase Excel status strings → AssetStatus enum value

STATUS_MAP: dict[str, AssetStatus] = {
    # Generic English
    "active":       AssetStatus.IN_USE,
    "in use":       AssetStatus.IN_USE,
    "in_use":       AssetStatus.IN_USE,
    "inuse":        AssetStatus.IN_USE,
    "available":    AssetStatus.AVAILABLE,
    "libre":        AssetStatus.AVAILABLE,
    "maintenance":  AssetStatus.MAINTENANCE,
    "retired":      AssetStatus.RETIRED,
    "inactive":     AssetStatus.RETIRED,
    "unknown":      AssetStatus.UNKNOWN,
    # Spanish
    "activo":       AssetStatus.IN_USE,
    "en uso":       AssetStatus.IN_USE,
    "disponible":   AssetStatus.AVAILABLE,
    "mantenimiento":AssetStatus.MAINTENANCE,
    "retirado":     AssetStatus.RETIRED,
    "baja":         AssetStatus.RETIRED,
    "inactivo":     AssetStatus.RETIRED,
    "desconocido":  AssetStatus.UNKNOWN,
    # ── Real TAT values (Sheet1, Mayo 2025) ───────────────────────────────────
    "in":           AssetStatus.AVAILABLE,    # in warehouse → available
    "dirty":        AssetStatus.MAINTENANCE,  # needs cleaning → maintenance
    "field":        AssetStatus.IN_USE,       # deployed in the field
    "in service":   AssetStatus.IN_USE,       # actively in service
}


# ── Cell helpers ───────────────────────────────────────────────────────────────

def _str(value: Any, max_len: int = MAX_FIELD_LEN) -> str | None:
    """Coerce cell value to a clean string, or None if blank."""
    if value is None:
        return None
    s = re.sub(r"\s+", " ", str(value).strip())
    return s[:max_len] if s else None


def _normalise_status(value: Any) -> AssetStatus:
    v = _str(value)
    return STATUS_MAP.get(v.lower() if v else "", AssetStatus.UNKNOWN)


def _meaningful_size(raw: str | None) -> str | None:
    """Return size string only when it carries real information."""
    if not raw:
        return None
    if raw.strip().lower() in SIZE_IGNORE_VALUES:
        return None
    return raw.strip()


# ── Column resolver ────────────────────────────────────────────────────────────

def _resolve_columns(
    headers: list[str],
) -> tuple[dict[str, int], ColumnMapping, list[str]]:
    normalised = [h.lower().strip() if h else "" for h in headers]
    col_map: dict[str, int] = {}

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            try:
                col_map[field] = normalised.index(alias)
                break
            except ValueError:
                continue

    detected = ColumnMapping(**{
        field: headers[idx]
        for field, idx in col_map.items()
        # only expose fields that are in ColumnMapping
        if field in ColumnMapping.model_fields and idx < len(headers)
    })

    matched_indices = set(col_map.values())
    unrecognised = [
        h for i, h in enumerate(headers)
        if i not in matched_indices and h and h.strip()
    ]

    return col_map, detected, unrecognised


def _get(row: tuple, col_map: dict[str, int], field: str) -> Any:
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# ── Part upsert ────────────────────────────────────────────────────────────────

def _upsert_part(
    db: Session,
    part_number: str,
    description: str | None,
    cache: dict[str, Part],
) -> tuple[Part, bool]:
    key = part_number.upper()
    if key in cache:
        return cache[key], False

    part = db.query(Part).filter(Part.part_number == part_number).first()
    if part:
        cache[key] = part
        return part, False

    part = Part(part_number=part_number, description=description)
    db.add(part)
    db.flush()
    cache[key] = part
    return part, True


# ── Asset upsert ───────────────────────────────────────────────────────────────

def _upsert_asset(
    db: Session,
    part: Part,
    serial_number: str | None,
    internal_code: str | None,
    item_name: str,
    status: AssetStatus,
    location: str | None,
    sn_cache: dict[str, Asset],
    ic_cache: dict[str, Asset],
) -> tuple[Asset, str]:
    existing: Asset | None = None

    if serial_number:
        existing = sn_cache.get(serial_number.upper()) or (
            db.query(Asset).filter(Asset.serial_number == serial_number).first()
        )
    if not existing and internal_code:
        existing = ic_cache.get(internal_code.upper()) or (
            db.query(Asset).filter(Asset.internal_code == internal_code).first()
        )

    if existing:
        changed = False
        updates: dict[str, Any] = {
            "item_name": item_name,
            "part_id":   part.id,
        }
        if status != AssetStatus.UNKNOWN:
            updates["status"] = status
        if location:
            updates["location"] = location

        for attr, new_val in updates.items():
            if getattr(existing, attr) != new_val:
                setattr(existing, attr, new_val)
                changed = True
        if changed:
            db.flush()
        return existing, "updated" if changed else "skipped"

    asset = Asset(
        part_id=part.id,
        serial_number=serial_number,
        internal_code=internal_code,
        item_name=item_name,
        status=status,
        location=location,
    )
    db.add(asset)
    db.flush()

    if serial_number:
        sn_cache[serial_number.upper()] = asset
    if internal_code:
        ic_cache[internal_code.upper()] = asset

    return asset, "created"


# ── Main public function ───────────────────────────────────────────────────────

def import_excel(
    db: Session,
    file_bytes: bytes,
    filename: str,
    sheet_name: str | None = None,
) -> ImportResult:
    """
    Parse a .xlsx file and upsert Parts + Assets into the database.
    Returns a detailed ImportResult. Raises ValueError on unreadable
    files; RuntimeError if the final commit fails.
    """
    # ── Open workbook ──────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo abrir el archivo Excel: {exc}") from exc

    available_sheets = wb.sheetnames

    if sheet_name:
        if sheet_name not in available_sheets:
            raise ValueError(
                f"Hoja '{sheet_name}' no encontrada. "
                f"Hojas disponibles: {', '.join(available_sheets)}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    actual_sheet = ws.title
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return ImportResult(filename=filename, sheet=actual_sheet,
                            available_sheets=available_sheets)

    # ── Resolve columns ────────────────────────────────────────────────────────
    header_raw = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    col_map, detected_columns, unrecognised = _resolve_columns(header_raw)
    data_rows = all_rows[1:]
    if len(data_rows) > MAX_ROWS_PER_IMPORT:
        data_rows = data_rows[:MAX_ROWS_PER_IMPORT]

    result = ImportResult(
        filename=filename,
        sheet=actual_sheet,
        available_sheets=available_sheets,
        detected_columns=detected_columns,
        unrecognised_columns=unrecognised,
        total_rows=0,
    )

    part_cache: dict[str, Part]  = {}
    sn_cache:   dict[str, Asset] = {}
    ic_cache:   dict[str, Asset] = {}

    # ── Process each row ───────────────────────────────────────────────────────
    for raw_idx, row in enumerate(data_rows):
        excel_row = raw_idx + 2

        if all(c is None or str(c).strip() == "" for c in row):
            continue

        result.total_rows += 1

        if len(result.errors) >= MAX_ERRORS_PER_IMPORT:
            break

        # Extract & normalise
        raw_pn    = _str(_get(row, col_map, "part_number"))
        raw_desc  = _str(_get(row, col_map, "part_desc"))    # PartDescription
        raw_sn    = _str(_get(row, col_map, "serial_number"))
        raw_ic    = _str(_get(row, col_map, "internal_code"))
        raw_name  = _str(_get(row, col_map, "item_name"))
        raw_size  = _str(_get(row, col_map, "size"))
        raw_loc   = _str(_get(row, col_map, "location"))
        raw_stat  = _get(row, col_map, "status")
        raw_ser   = _str(_get(row, col_map, "series"))

        # Normalise part_number: uppercase, collapse all whitespace
        part_number = re.sub(r"\s+", "", raw_pn).upper() if raw_pn else None

        # Identifiers — uppercase for consistent dedup
        serial_number = raw_sn.upper() if raw_sn else None
        internal_code = raw_ic.upper() if raw_ic else None

        # When no serial AND no internal_code, try using Series as fallback
        # internal_code (only if Series looks like a meaningful code, not a
        # pure integer like "702" which is just a product family number)
        if not serial_number and not internal_code and raw_ser:
            if not raw_ser.isdigit():           # skip pure numeric series
                internal_code = f"SER-{raw_ser.upper()}"

        # item_name: prefer PartDescription if available, otherwise Item Name
        item_name = raw_desc or raw_name or "Sin nombre"
        size_str  = _meaningful_size(raw_size)
        if size_str and size_str.lower() not in item_name.lower():
            item_name = f"{item_name} ({size_str})"
        item_name = item_name[:MAX_FIELD_LEN]

        status   = _normalise_status(raw_stat)
        location = raw_loc[:MAX_FIELD_LEN] if raw_loc else None
        identifier = serial_number or internal_code

        # ── Validate ───────────────────────────────────────────────────────────
        if not part_number:
            result.errors.append(ImportRowError(
                row=excel_row, identifier=identifier,
                reason="'PartNumber' vacío o columna no detectada.",
            ))
            result.rows_skipped += 1
            continue

        if not serial_number and not internal_code:
            result.errors.append(ImportRowError(
                row=excel_row, identifier=None,
                reason=(
                    "Se requiere 'Serial Number' o 'Internal Code'. "
                    "Ambos están vacíos (Series numérico ignorado como código)."
                ),
            ))
            result.rows_skipped += 1
            continue

        # ── Upsert Part ────────────────────────────────────────────────────────
        sp_part = db.begin_nested()
        try:
            # Use PartDescription as Part.description if available
            part_desc = raw_desc or raw_name
            part, part_created = _upsert_part(db, part_number, part_desc, part_cache)
        except Exception as exc:
            sp_part.rollback()
            result.errors.append(ImportRowError(
                row=excel_row, identifier=identifier,
                reason=f"Error al crear Part '{part_number}': {exc}",
            ))
            result.rows_skipped += 1
            continue
        else:
            sp_part.commit()
            if part_created:
                result.parts_created += 1
            else:
                result.parts_reused += 1

        # ── Upsert Asset ───────────────────────────────────────────────────────
        sp_asset = db.begin_nested()
        try:
            _asset, action = _upsert_asset(
                db, part,
                serial_number=serial_number,
                internal_code=internal_code,
                item_name=item_name,
                status=status,
                location=location,
                sn_cache=sn_cache,
                ic_cache=ic_cache,
            )
        except IntegrityError as exc:
            sp_asset.rollback()
            result.errors.append(ImportRowError(
                row=excel_row, identifier=identifier,
                reason=f"Conflicto de unicidad: {exc.orig}",
            ))
            result.rows_skipped += 1
            continue
        except Exception as exc:
            sp_asset.rollback()
            result.errors.append(ImportRowError(
                row=excel_row, identifier=identifier,
                reason=f"Error inesperado: {exc}",
            ))
            result.rows_skipped += 1
            continue
        else:
            sp_asset.commit()
            if action == "created":
                result.assets_created += 1
            elif action == "updated":
                result.assets_updated += 1
            else:
                result.rows_skipped += 1

    # ── Final commit ───────────────────────────────────────────────────────────
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise RuntimeError(
            f"Error al confirmar la importación: {exc}"
        ) from exc

    return result
